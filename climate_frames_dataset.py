"""
Shared dataset utilities for the Climate Frames workspace.

This module centralizes:
  - the default workbook path
  - row filtering / normalization
  - exact-text duplicate handling
  - token summary parsing
  - trigger-token span matching for BIO conversion
"""

from __future__ import annotations

import os
import re
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd


DEFAULT_DATA_PATH = Path(
    os.environ.get(
        "CLFRAMES_DATA_PATH",
        r"E:\Frames\newdata\14 May 2026 12 articles Ann. Core Peripheral RST and FrameNET Structure.xlsx",
    )
)
ANNOTATION_SHEET = "Core and Peripheral Annotations"
TOKEN_SUMMARY_SHEET = "Token Summary"

_LIST_SPLIT_RE = re.compile(r"\s*[;,]\s*")
_WS_RE = re.compile(r"\s+")


def resolve_data_path(path: str | Path | None = None) -> Path:
    return Path(path) if path else DEFAULT_DATA_PATH


def normalize_frame(name: str | None) -> str:
    if name is None:
        return ""
    text = _WS_RE.sub(" ", str(name).replace("_", " ").strip())
    return text


def normalize_peripheral_frame(name: str | None) -> str:
    return normalize_frame(name).lower()


def normalize_text(text: str | None) -> str:
    if text is None:
        return ""
    return str(text).strip()


def split_annotation_items(value: str | None) -> list[str]:
    if value is None:
        return []
    raw = str(value).strip()
    if not raw or raw.lower() == "nan":
        return []
    return [part.strip() for part in _LIST_SPLIT_RE.split(raw) if part.strip()]


def dedupe_preserve_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        ordered.append(value)
    return ordered


def is_valid_annotation_row(text: str, core_frame: str) -> bool:
    if not text or not core_frame:
        return False
    if core_frame.lower() in {"none", "nan"}:
        return False
    if core_frame.startswith("(This is a section heading"):
        return False
    return True


def _parse_annotation_rows(path: str | Path | None = None) -> list[dict]:
    workbook_path = resolve_data_path(path)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[ANNOTATION_SHEET]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers)}

    required = [
        "Text Segment",
        "Core Frame",
        "Peripheral Frames",
        "Trigger Tokens",
        "Frame Roles",
    ]
    missing = [name for name in required if name not in col]
    if missing:
        raise ValueError(f"Missing expected columns in {workbook_path.name}: {missing}")

    rows: list[dict] = []
    for rownum, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue

        text = normalize_text(values[col["Text Segment"]])
        core_frame = normalize_frame(values[col["Core Frame"]])
        if not is_valid_annotation_row(text, core_frame):
            continue

        peripheral_frames = dedupe_preserve_order(
            normalize_peripheral_frame(item) for item in split_annotation_items(values[col["Peripheral Frames"]])
        )
        trigger_tokens = dedupe_preserve_order(
            item.lower() for item in split_annotation_items(values[col["Trigger Tokens"]]) if item.lower() != "nan"
        )
        frame_roles = normalize_text(values[col["Frame Roles"]])

        rows.append(
            {
                "row_id": rownum,
                "text": text,
                "core_frame": core_frame,
                "peripheral_frames": peripheral_frames,
                "all_frames": [core_frame] + peripheral_frames,
                "tokens": trigger_tokens,
                "frame_roles": frame_roles,
            }
        )
    return rows


def load_annotations(
    path: str | Path | None = None,
    dedupe_mode: str = "merge",
) -> pd.DataFrame:
    """
    Load workbook annotations.

    dedupe_mode:
      - "none": keep every row
      - "first": keep the first row for each exact text
      - "merge": merge repeated exact texts by unioning peripherals/tokens
    """

    rows = _parse_annotation_rows(path)
    if dedupe_mode == "none":
        return pd.DataFrame(rows)

    grouped: OrderedDict[str, list[dict]] = OrderedDict()
    for row in rows:
        grouped.setdefault(row["text"], []).append(row)

    merged_rows: list[dict] = []
    for text, variants in grouped.items():
        base = dict(variants[0])
        base["duplicate_count"] = len(variants)
        base["source_rows"] = [item["row_id"] for item in variants]

        unique_core = {item["core_frame"] for item in variants}
        if len(unique_core) != 1:
            raise ValueError(f"Conflicting core labels for duplicated text: {text[:120]!r}")

        if dedupe_mode == "first":
            merged_rows.append(base)
            continue

        peripheral_union: list[str] = []
        token_union: list[str] = []
        role_variants: list[str] = []
        for item in variants:
            peripheral_union.extend(item["peripheral_frames"])
            token_union.extend(item["tokens"])
            if item["frame_roles"]:
                role_variants.append(item["frame_roles"])

        base["peripheral_frames"] = dedupe_preserve_order(peripheral_union)
        base["tokens"] = dedupe_preserve_order(token_union)
        base["all_frames"] = [base["core_frame"]] + base["peripheral_frames"]
        base["role_variants"] = dedupe_preserve_order(role_variants)
        base["frame_roles"] = " || ".join(base["role_variants"])
        merged_rows.append(base)

    return pd.DataFrame(merged_rows)


def load_token_summary(path: str | Path | None = None) -> pd.DataFrame:
    workbook_path = resolve_data_path(path)
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if TOKEN_SUMMARY_SHEET not in wb.sheetnames:
        return pd.DataFrame(columns=["frame_type", "frame_name", "evoking_tokens"])

    ws = wb[TOKEN_SUMMARY_SHEET]
    records: list[dict] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        frame_type = normalize_frame(values[0]) if len(values) > 0 else ""
        frame_name = normalize_frame(values[1]) if len(values) > 1 else ""
        tokens = split_annotation_items(values[2] if len(values) > 2 else None)
        evoking_tokens = dedupe_preserve_order(
            token.lower() for token in tokens if token.lower() != "nan"
        )
        if not frame_name:
            continue
        records.append(
            {
                "frame_type": frame_type,
                "frame_name": frame_name,
                "evoking_tokens": evoking_tokens,
            }
        )
    return pd.DataFrame(records)


def token_summary_by_frame(path: str | Path | None = None) -> dict[str, list[str]]:
    summary = load_token_summary(path)
    grouped: dict[str, list[str]] = defaultdict(list)
    for _, row in summary.iterrows():
        grouped[row["frame_name"]].extend(row["evoking_tokens"])
    return {frame: dedupe_preserve_order(tokens) for frame, tokens in grouped.items()}


def find_token_spans(text: str, tokens: Iterable[str]) -> list[tuple[int, int, str]]:
    """
    Find non-overlapping character spans for trigger tokens.

    Matching is case-insensitive and greedy by longest span first.
    """

    candidates: list[tuple[int, int, str]] = []
    for token in dedupe_preserve_order(token for token in tokens if token):
        pattern = re.escape(token)
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            candidates.append((match.start(), match.end(), token))

    candidates.sort(key=lambda item: (item[0], -(item[1] - item[0]), item[2]))

    accepted: list[tuple[int, int, str]] = []
    occupied: list[tuple[int, int]] = []
    for start, end, token in candidates:
        overlaps = any(not (end <= occ_start or start >= occ_end) for occ_start, occ_end in occupied)
        if overlaps:
            continue
        accepted.append((start, end, token))
        occupied.append((start, end))

    return accepted


def dataset_profile(df: pd.DataFrame) -> dict:
    return {
        "rows": int(len(df)),
        "unique_core_frames": int(df["core_frame"].nunique()) if not df.empty else 0,
        "core_counts": df["core_frame"].value_counts().to_dict() if not df.empty else {},
        "texts_with_tokens": int(df["tokens"].map(bool).sum()) if not df.empty else 0,
    }
