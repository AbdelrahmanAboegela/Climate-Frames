"""
ClimateBERT Base Model POC — Round 2: Additional Tests
======================================================
1. Zero-shot frame classification (cosine sim: paragraph vs frame name)
2. Intra-frame vs inter-frame cosine similarity
3. Frame name embedding analysis (frame taxonomy visualization)
4. Tokenizer coverage analysis (subword split quality)

All outputs saved to e:/Frames/poc_outputs/
"""

import os, re, warnings, json
import numpy as np
import pandas as pd
import torch
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import seaborn as sns
from transformers import AutoTokenizer, AutoModel, logging as hf_logging
from sklearn.metrics.pairwise import cosine_similarity

hf_logging.set_verbosity_error()
warnings.filterwarnings("ignore")

OUTPUT_DIR = r"e:\Frames\poc_outputs"
MODEL_NAME = "climatebert/distilroberta-base-climate-f"
DATA_PATH = r"e:\Frames\3 Articles Samples Annotation 2026.xlsx"


def normalize_frame(name: str) -> str:
    name = name.strip()
    name = re.sub(r"\s*_\s*", "_", name)
    name = re.sub(r"\s+", " ", name)
    return name


def load_data(path: str) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        text, core, periph, tokens = row
        if not text:
            continue
        core_frame = normalize_frame(str(core)) if core else ""
        periph_frames = [normalize_frame(p) for p in str(periph).split(",") if p.strip()] if periph else []
        all_frames = [core_frame] + periph_frames
        token_list = [t.strip().lower() for t in str(tokens).split(",") if t.strip()] if tokens else []
        records.append({
            "text": str(text).strip(),
            "core_frame": core_frame,
            "peripheral_frames": periph_frames,
            "all_frames": all_frames,
            "tokens": token_list,
        })
    return pd.DataFrame(records)


def embed_texts(texts: list[str], tokenizer, model, device) -> np.ndarray:
    """Mean-pooled embeddings."""
    embeddings = []
    model.eval()
    with torch.no_grad():
        for text in texts:
            inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs)
            mask = inputs["attention_mask"].unsqueeze(-1)
            hidden = outputs.last_hidden_state
            mean_emb = (hidden * mask).sum(dim=1) / mask.sum(dim=1)
            embeddings.append(mean_emb.cpu().numpy().squeeze())
    return np.array(embeddings)


# ═══════════════════════════════════════════════════════════
# TEST 1: Zero-Shot Frame Classification
# ═══════════════════════════════════════════════════════════

def zero_shot_classification(df, tokenizer, model, device):
    """Classify each paragraph by cosine similarity to frame name embeddings."""
    print("\n" + "─" * 55)
    print("🔬 Test 1: Zero-Shot Frame Classification")
    print("─" * 55)

    # Collect all unique frames (core + peripheral)
    all_frames = sorted(set(df["core_frame"].tolist()))

    # Embed all paragraphs
    print("  Embedding paragraphs...")
    para_embs = embed_texts(df["text"].tolist(), tokenizer, model, device)

    # Embed all frame names (treat frame name as a short text)
    # Use more descriptive prompts for better matching
    frame_descriptions = [f"This text is about {f.replace('_', ' ').lower()}" for f in all_frames]
    print(f"  Embedding {len(all_frames)} frame names...")
    frame_embs = embed_texts(frame_descriptions, tokenizer, model, device)

    # Cosine similarity matrix: paragraphs × frames
    sim_matrix = cosine_similarity(para_embs, frame_embs)

    # For each paragraph, find predicted frame (highest cosine sim)
    correct = 0
    top3_correct = 0
    results = []
    for i, row in df.iterrows():
        true_frame = row["core_frame"]
        sims = sim_matrix[i]
        ranked_indices = np.argsort(sims)[::-1]
        ranked_frames = [all_frames[j] for j in ranked_indices]
        predicted = ranked_frames[0]
        top3 = ranked_frames[:3]

        is_correct = predicted == true_frame
        is_top3 = true_frame in top3
        if is_correct:
            correct += 1
        if is_top3:
            top3_correct += 1

        results.append({
            "text_preview": row["text"][:60] + "...",
            "true_frame": true_frame,
            "predicted_frame": predicted,
            "correct": is_correct,
            "top3": top3,
            "top3_correct": is_top3,
            "confidence": float(sims[ranked_indices[0]]),
        })

    acc = correct / len(df) * 100
    top3_acc = top3_correct / len(df) * 100

    print(f"\n  📊 Zero-Shot Results:")
    print(f"     Top-1 Accuracy: {correct}/{len(df)} = {acc:.1f}%")
    print(f"     Top-3 Accuracy: {top3_correct}/{len(df)} = {top3_acc:.1f}%")
    print(f"\n  Sample predictions:")
    for r in results[:8]:
        mark = "✓" if r["correct"] else "✗"
        print(f"    {mark} True: {r['true_frame']:30s} → Predicted: {r['predicted_frame']:30s} (sim={r['confidence']:.3f})")

    # Save detailed results
    with open(os.path.join(OUTPUT_DIR, "zero_shot_results.txt"), "w", encoding="utf-8") as f:
        f.write(f"Zero-Shot Frame Classification — ClimateBERT Base Model\n")
        f.write(f"{'='*55}\n")
        f.write(f"Top-1 Accuracy: {acc:.1f}%\n")
        f.write(f"Top-3 Accuracy: {top3_acc:.1f}%\n\n")
        for r in results:
            mark = "✓" if r["correct"] else "✗"
            f.write(f"{mark} True: {r['true_frame']:35s} Pred: {r['predicted_frame']:35s} sim={r['confidence']:.3f}\n")
            f.write(f"  Top-3: {r['top3']}\n\n")
    print("  ✓ Saved zero_shot_results.txt")

    return acc, top3_acc, sim_matrix, all_frames


# ═══════════════════════════════════════════════════════════
# TEST 2: Intra-Frame vs Inter-Frame Similarity
# ═══════════════════════════════════════════════════════════

def intra_inter_similarity(df, tokenizer, model, device):
    """Compare cosine similarity within same frame vs across different frames."""
    print("\n" + "─" * 55)
    print("🔬 Test 2: Intra-Frame vs Inter-Frame Cosine Similarity")
    print("─" * 55)

    para_embs = embed_texts(df["text"].tolist(), tokenizer, model, device)
    sim_matrix = cosine_similarity(para_embs)

    # Find frames with >= 2 paragraphs
    frame_counts = df["core_frame"].value_counts()
    multi_frames = frame_counts[frame_counts >= 2].index.tolist()
    print(f"  Frames with ≥2 paragraphs: {len(multi_frames)}")
    for f in multi_frames:
        print(f"    {f}: {frame_counts[f]} paragraphs")

    intra_sims = []
    inter_sims = []

    for i in range(len(df)):
        for j in range(i + 1, len(df)):
            s = sim_matrix[i, j]
            if df.iloc[i]["core_frame"] == df.iloc[j]["core_frame"]:
                intra_sims.append(s)
            else:
                inter_sims.append(s)

    # Also check peripheral frame overlap
    periph_overlap_sims = []
    no_overlap_sims = []
    for i in range(len(df)):
        for j in range(i + 1, len(df)):
            s = sim_matrix[i, j]
            frames_i = set(df.iloc[i]["all_frames"])
            frames_j = set(df.iloc[j]["all_frames"])
            if frames_i & frames_j:  # any shared frame
                periph_overlap_sims.append(s)
            else:
                no_overlap_sims.append(s)

    print(f"\n  📊 Core Frame Similarity:")
    print(f"     Intra-frame (same core):  mean={np.mean(intra_sims):.4f}, std={np.std(intra_sims):.4f}, n={len(intra_sims)}")
    print(f"     Inter-frame (diff core):  mean={np.mean(inter_sims):.4f}, std={np.std(inter_sims):.4f}, n={len(inter_sims)}")
    print(f"     Gap (intra - inter):      {np.mean(intra_sims) - np.mean(inter_sims):.4f}")

    print(f"\n  📊 Any Frame Overlap:")
    print(f"     Overlapping frames:       mean={np.mean(periph_overlap_sims):.4f}, n={len(periph_overlap_sims)}")
    print(f"     No overlap:               mean={np.mean(no_overlap_sims):.4f}, n={len(no_overlap_sims)}")
    print(f"     Gap:                      {np.mean(periph_overlap_sims) - np.mean(no_overlap_sims):.4f}")

    # Plot distribution
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    axes[0].hist(intra_sims, bins=15, alpha=0.7, label=f"Same Frame (n={len(intra_sims)})", color="#2ecc71", edgecolor="white")
    axes[0].hist(inter_sims, bins=25, alpha=0.5, label=f"Different Frame (n={len(inter_sims)})", color="#e74c3c", edgecolor="white")
    axes[0].axvline(np.mean(intra_sims), color="#27ae60", linestyle="--", linewidth=2, label=f"Intra mean={np.mean(intra_sims):.3f}")
    axes[0].axvline(np.mean(inter_sims), color="#c0392b", linestyle="--", linewidth=2, label=f"Inter mean={np.mean(inter_sims):.3f}")
    axes[0].set_title("Core Frame: Intra vs Inter Similarity", fontweight="bold")
    axes[0].set_xlabel("Cosine Similarity")
    axes[0].legend(fontsize=8)

    axes[1].hist(periph_overlap_sims, bins=20, alpha=0.7, label=f"Frame Overlap (n={len(periph_overlap_sims)})", color="#3498db", edgecolor="white")
    axes[1].hist(no_overlap_sims, bins=20, alpha=0.5, label=f"No Overlap (n={len(no_overlap_sims)})", color="#e67e22", edgecolor="white")
    axes[1].axvline(np.mean(periph_overlap_sims), color="#2980b9", linestyle="--", linewidth=2)
    axes[1].axvline(np.mean(no_overlap_sims), color="#d35400", linestyle="--", linewidth=2)
    axes[1].set_title("Any Frame Overlap vs No Overlap", fontweight="bold")
    axes[1].set_xlabel("Cosine Similarity")
    axes[1].legend(fontsize=8)

    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "intra_inter_similarity.png"), dpi=200)
    plt.close()
    print("  ✓ Saved intra_inter_similarity.png")

    return np.mean(intra_sims), np.mean(inter_sims)


# ═══════════════════════════════════════════════════════════
# TEST 3: Frame Name Embedding Analysis
# ═══════════════════════════════════════════════════════════

def frame_taxonomy_analysis(df, tokenizer, model, device):
    """Embed frame names and visualize their relationships."""
    print("\n" + "─" * 55)
    print("🔬 Test 3: Frame Name Taxonomy (Model's View)")
    print("─" * 55)

    # Collect ALL unique frames (core + peripheral)
    all_frames_set = set()
    for _, row in df.iterrows():
        all_frames_set.add(row["core_frame"])
        all_frames_set.update(row["peripheral_frames"])
    all_frames = sorted(all_frames_set)
    print(f"  Total unique frames (core + peripheral): {len(all_frames)}")

    # Embed frame names
    frame_texts = [f.replace("_", " ") for f in all_frames]
    frame_embs = embed_texts(frame_texts, tokenizer, model, device)

    # Cosine similarity
    sim = cosine_similarity(frame_embs)

    # Heatmap
    fig, ax = plt.subplots(figsize=(22, 18))
    sns.heatmap(pd.DataFrame(sim, index=all_frames, columns=all_frames),
                annot=True, fmt=".2f", cmap="RdYlBu_r", center=0.5,
                annot_kws={"fontsize": 4.5}, linewidths=0.2, ax=ax)
    ax.set_title("All Frames Similarity (Core + Peripheral) — ClimateBERT's View", fontsize=12, fontweight="bold")
    ax.tick_params(labelsize=5.5)
    plt.xticks(rotation=45, ha="right")
    plt.yticks(rotation=0)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "frame_taxonomy_heatmap.png"), dpi=200, bbox_inches="tight")
    plt.close()
    print("  ✓ Saved frame_taxonomy_heatmap.png")

    # UMAP of frame names
    try:
        import umap
    except ImportError:
        import umap.umap_ as umap

    if len(all_frames) > 5:
        reducer = umap.UMAP(n_neighbors=min(8, len(all_frames) - 1), min_dist=0.3, random_state=42, metric="cosine")
        coords = reducer.fit_transform(frame_embs)

        # Color: core frames in blue, peripheral-only in orange
        core_set = set(df["core_frame"].tolist())
        colors = ["#2980b9" if f in core_set else "#e67e22" for f in all_frames]

        fig, ax = plt.subplots(figsize=(16, 12))
        ax.scatter(coords[:, 0], coords[:, 1], c=colors, s=150, edgecolors="white", linewidth=1, alpha=0.85)
        for i, f in enumerate(all_frames):
            ax.annotate(f, (coords[i, 0], coords[i, 1]), fontsize=6, ha="center", va="bottom", alpha=0.8)

        from matplotlib.patches import Patch
        legend_elements = [Patch(facecolor="#2980b9", label="Core Frame"),
                           Patch(facecolor="#e67e22", label="Peripheral Only")]
        ax.legend(handles=legend_elements, fontsize=9)
        ax.set_title("Frame Taxonomy UMAP — How ClimateBERT Sees Frame Relations", fontsize=13, fontweight="bold")
        plt.tight_layout()
        plt.savefig(os.path.join(OUTPUT_DIR, "frame_taxonomy_umap.png"), dpi=200, bbox_inches="tight")
        plt.close()
        print("  ✓ Saved frame_taxonomy_umap.png")

    # Find most similar frame pairs
    print("\n  Top 10 most similar frame pairs:")
    pairs = []
    for i in range(len(all_frames)):
        for j in range(i + 1, len(all_frames)):
            pairs.append((all_frames[i], all_frames[j], sim[i, j]))
    pairs.sort(key=lambda x: x[2], reverse=True)
    for a, b, s in pairs[:10]:
        print(f"    {s:.3f}  {a} ↔ {b}")

    return all_frames


# ═══════════════════════════════════════════════════════════
# TEST 4: Tokenizer Coverage Analysis
# ═══════════════════════════════════════════════════════════

def tokenizer_coverage(df, tokenizer):
    """Analyze how ClimateBERT's tokenizer handles frame-evoking tokens."""
    print("\n" + "─" * 55)
    print("🔬 Test 4: Tokenizer Coverage Analysis")
    print("─" * 55)

    results = []
    single_token = 0
    multi_token = 0
    total = 0

    for _, row in df.iterrows():
        for tok_str in row["tokens"]:
            encoded = tokenizer.encode(tok_str, add_special_tokens=False)
            decoded_parts = [tokenizer.decode([t]).strip() for t in encoded]
            n_subwords = len(encoded)
            total += 1
            if n_subwords == 1:
                single_token += 1
            else:
                multi_token += 1
            results.append({
                "token": tok_str,
                "subwords": n_subwords,
                "parts": decoded_parts,
                "frame": row["core_frame"],
            })

    print(f"\n  📊 Tokenizer Stats:")
    print(f"     Total frame-evoking tokens: {total}")
    print(f"     Single subword (kept whole): {single_token} ({single_token/total*100:.1f}%)")
    print(f"     Multi-subword (split):       {multi_token} ({multi_token/total*100:.1f}%)")

    # Show examples of heavily split tokens
    results.sort(key=lambda x: x["subwords"], reverse=True)
    print(f"\n  Most split tokens:")
    for r in results[:10]:
        print(f"    \"{r['token']}\" → {r['subwords']} subwords: {r['parts']}")

    # Distribution
    subword_counts = [r["subwords"] for r in results]
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.hist(subword_counts, bins=range(1, max(subword_counts) + 2), edgecolor="white", color="#3498db", alpha=0.8, align="left")
    ax.set_xlabel("Number of Subword Tokens")
    ax.set_ylabel("Frequency")
    ax.set_title("Frame-Evoking Token Subword Split Distribution", fontweight="bold")
    ax.set_xticks(range(1, max(subword_counts) + 1))
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "tokenizer_coverage.png"), dpi=200)
    plt.close()
    print("  ✓ Saved tokenizer_coverage.png")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  ClimateBERT POC — Round 2: Additional Tests")
    print("=" * 60)

    df = load_data(DATA_PATH)
    print(f"\n📂 Loaded {len(df)} paragraphs")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"🤖 Loading model on {device}...")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    model = AutoModel.from_pretrained(MODEL_NAME).to(device)
    print("   ✓ Ready\n")

    # Run all tests
    zs_acc, zs_top3, _, _ = zero_shot_classification(df, tokenizer, model, device)
    intra_mean, inter_mean = intra_inter_similarity(df, tokenizer, model, device)
    frame_taxonomy_analysis(df, tokenizer, model, device)
    tokenizer_coverage(df, tokenizer)

    # Summary
    print("\n" + "=" * 60)
    print("  📋 Round 2 Summary")
    print("=" * 60)
    print(f"  Zero-shot Top-1 Accuracy: {zs_acc:.1f}%")
    print(f"  Zero-shot Top-3 Accuracy: {zs_top3:.1f}%")
    print(f"  Intra-frame similarity:   {intra_mean:.4f}")
    print(f"  Inter-frame similarity:   {inter_mean:.4f}")
    print(f"  Similarity gap:           {intra_mean - inter_mean:.4f}")
    print(f"\n  All outputs → {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
