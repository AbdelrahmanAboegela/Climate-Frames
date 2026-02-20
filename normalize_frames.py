"""
Frame Name Normalization Script
================================
Reads the annotation Excel file and normalizes frame names:
1. Fixes spacing/underscore inconsistencies
2. Reports duplicate and near-duplicate frames
3. Saves a cleaned version of the spreadsheet

Run: python normalize_frames.py
Output: e:\Frames\3 Articles Samples Annotation 2026 - Normalized.xlsx
"""

import re
import openpyxl
from collections import Counter

INPUT_PATH = r"e:\Frames\3 Articles Samples Annotation 2026.xlsx"
OUTPUT_PATH = r"e:\Frames\3 Articles Samples Annotation 2026 - Normalized.xlsx"


# ── Normalization Rules ──────────────────────────────────
# Map inconsistent names → canonical name
# Add more mappings here after linguistics team review

FRAME_MERGE_MAP = {
    # Underscore vs space duplicates (auto-detected)
    "Emission_Generation": "Emission Generation",
    "Human_Impact": "Human Impact",
    "Danger_Threat": "Danger Threat",
    "Climate_Action": "Climate Action",
    "Forced_Migration": "Forced Migration",
    "Energy_Transition": "Energy Transition",
    "Climate_Risk_Escalation": "Climate Risk Escalation",
    "Ecosystem_Degradation": "Ecosystem Degradation",

    # Near-duplicates needing review (uncomment after confirmation):
    # "Collective Climate Action": "Collective Action",
    # "Disaster Intensification_Threat": "Disaster Intensification",
}


def normalize_frame(name: str) -> str:
    """Normalize a single frame name."""
    name = name.strip()

    # First, check explicit merge map
    if name in FRAME_MERGE_MAP:
        return FRAME_MERGE_MAP[name]

    # Replace underscores with spaces (general rule)
    name = name.replace("_", " ")

    # Clean up extra whitespace
    name = re.sub(r"\s+", " ", name).strip()

    # Title case for consistency
    name = name.title()

    return name


def normalize_frame_list(frame_str: str) -> str:
    """Normalize a comma-separated list of frames."""
    if not frame_str or str(frame_str).strip() in ("None", "nan", ""):
        return ""
    frames = [normalize_frame(f.strip()) for f in str(frame_str).split(",") if f.strip()]
    return ", ".join(frames)


def main():
    print("=" * 55)
    print("  Frame Name Normalization")
    print("=" * 55)

    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb["Sheet1"]

    # Collect all frames before normalization
    all_core_before = []
    all_periph_before = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        text, core, periph, tokens = row
        if not text:
            continue
        if core:
            all_core_before.append(str(core).strip())
        if periph:
            for p in str(periph).split(","):
                if p.strip():
                    all_periph_before.append(p.strip())

    all_before = set(all_core_before + all_periph_before)
    print(f"\n  Unique frames BEFORE normalization: {len(all_before)}")

    # Apply normalization
    all_core_after = []
    all_periph_after = []
    changes = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        text_cell, core_cell, periph_cell, tokens_cell = row

        if not text_cell.value:
            continue

        # Normalize core frame
        if core_cell.value:
            old = str(core_cell.value).strip()
            new = normalize_frame(old)
            if old != new:
                changes += 1
                print(f"  Core:  \"{old}\" → \"{new}\"")
            core_cell.value = new
            all_core_after.append(new)

        # Normalize peripheral frames
        if periph_cell.value:
            old = str(periph_cell.value).strip()
            new = normalize_frame_list(old)
            if old != new:
                changes += 1
                print(f"  Periph: \"{old}\" → \"{new}\"")
            periph_cell.value = new
            for p in new.split(","):
                if p.strip():
                    all_periph_after.append(p.strip())

    all_after = set(all_core_after + all_periph_after)
    print(f"\n  Unique frames AFTER normalization: {len(all_after)}")
    print(f"  Frames reduced by: {len(all_before) - len(all_after)}")
    print(f"  Total cell changes: {changes}")

    # Report remaining potential duplicates (similarity-based)
    print(f"\n  ── Remaining frames for review ──")
    sorted_frames = sorted(all_after)
    for i, f1 in enumerate(sorted_frames):
        for f2 in sorted_frames[i+1:]:
            # Check for near-duplicates (one is substring of another)
            if f1.lower() in f2.lower() or f2.lower() in f1.lower():
                if f1 != f2:
                    print(f"  ⚠ Similar: \"{f1}\" ↔ \"{f2}\"")

    # Frame frequency
    print(f"\n  ── Core Frame Frequency ──")
    core_counts = Counter(all_core_after)
    for frame, count in core_counts.most_common():
        marker = "⚠" if count < 2 else "✓"
        print(f"  {marker} {frame}: {count}")

    # Save
    wb.save(OUTPUT_PATH)
    print(f"\n  ✓ Saved normalized file to:")
    print(f"    {OUTPUT_PATH}")
    print("=" * 55)


if __name__ == "__main__":
    main()
