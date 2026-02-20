"""
ClimateBERT Base Model POC — Capabilities Visualization
========================================================
Probes the base `climatebert/distilroberta-base-climate-f` model
(pre-finetuning) to understand its representation quality for
climate frame classification and token extraction.

Modules:
  1. Paragraph-level embedding UMAP (Core Frame + All Frames views)
  2. Token-level embedding UMAP for frame-evoking tokens
  3. Frame-to-frame cosine similarity heatmap
  4. MLM probing (mask frame tokens, check predictions)
  5. Attention weight analysis on sample paragraphs

All outputs saved to e:/Frames/poc_outputs/
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
import torch
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import seaborn as sns
from transformers import AutoTokenizer, AutoModel, AutoModelForMaskedLM, logging as hf_logging
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.metrics import silhouette_score

hf_logging.set_verbosity_error()
warnings.filterwarnings("ignore")

OUTPUT_DIR = r"e:\Frames\poc_outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

MODEL_NAME = "climatebert/distilroberta-base-climate-f"
DATA_PATH = r"e:\Frames\3 Articles Samples Annotation 2026.xlsx"

# ═══════════════════════════════════════════════════════════
# PART 0: Data Loading & Preprocessing
# ═══════════════════════════════════════════════════════════

def normalize_frame(name: str) -> str:
    """Standardize frame names: strip, collapse spaces, unify separators."""
    name = name.strip()
    name = re.sub(r"\s*_\s*", "_", name)       # "Danger _ Threat" → "Danger_Threat"
    name = re.sub(r"\s+", " ", name)            # collapse whitespace
    return name


def load_data(path: str) -> pd.DataFrame:
    """Load annotated Excel and parse into structured DataFrame."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        text, core, periph, tokens = row
        if not text:
            continue
        core_frame = normalize_frame(str(core)) if core else ""
        periph_frames = [normalize_frame(p) for p in str(periph).split(",") if p.strip()] if periph else []
        all_frames = [core_frame] + periph_frames
        token_list = [t.strip().lower() for t in str(tokens).split(",") if t.strip()] if tokens else []
        records.append({
            "text": str(text).strip(),
            "core_frame": core_frame,
            "peripheral_frames": periph_frames,
            "all_frames": all_frames,
            "tokens": token_list,
        })
    return pd.DataFrame(records)


def find_token_positions(text: str, token_list: list[str]) -> list[dict]:
    """Find character start/end positions of frame-evoking tokens in text."""
    text_lower = text.lower()
    positions = []
    for token in token_list:
        tok_lower = token.lower()
        start = 0
        while True:
            idx = text_lower.find(tok_lower, start)
            if idx == -1:
                break
            positions.append({
                "token": token,
                "start": idx,
                "end": idx + len(token),
                "matched_text": text[idx:idx + len(token)],
            })
            break  # take first match only to avoid duplicates
    return positions


# ═══════════════════════════════════════════════════════════
# PART 1: Paragraph-Level Embedding UMAP
# ═══════════════════════════════════════════════════════════

def extract_paragraph_embeddings(texts: list[str], tokenizer, model, device) -> np.ndarray:
    """Extract [CLS] embeddings for each paragraph."""
    embeddings = []
    model.eval()
    with torch.no_grad():
        for text in texts:
            inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs)
            # Use [CLS] token (first token) representation
            cls_emb = outputs.last_hidden_state[:, 0, :].cpu().numpy().squeeze()
            embeddings.append(cls_emb)
    return np.array(embeddings)


def extract_mean_pooled_embeddings(texts: list[str], tokenizer, model, device) -> np.ndarray:
    """Extract mean-pooled embeddings (average all tokens) for each paragraph."""
    embeddings = []
    model.eval()
    with torch.no_grad():
        for text in texts:
            inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs)
            attention_mask = inputs["attention_mask"].unsqueeze(-1)
            hidden = outputs.last_hidden_state
            masked = hidden * attention_mask
            mean_emb = masked.sum(dim=1) / attention_mask.sum(dim=1)
            embeddings.append(mean_emb.cpu().numpy().squeeze())
    return np.array(embeddings)


def plot_umap_paragraphs(embeddings: np.ndarray, labels: list[str], title: str, filename: str):
    """UMAP 2D visualization of paragraph embeddings colored by frame label."""
    try:
        import umap
    except ImportError:
        import umap.umap_ as umap

    reducer = umap.UMAP(n_neighbors=min(10, len(embeddings) - 1), min_dist=0.3, random_state=42, metric="cosine")
    coords = reducer.fit_transform(embeddings)

    unique_labels = sorted(set(labels))
    color_map = {label: cm.tab20(i / max(len(unique_labels), 1)) for i, label in enumerate(unique_labels)}

    fig, ax = plt.subplots(figsize=(16, 12))
    for label in unique_labels:
        mask = [l == label for l in labels]
        xs = coords[mask, 0]
        ys = coords[mask, 1]
        ax.scatter(xs, ys, c=[color_map[label]], label=label, s=120, edgecolors="white", linewidth=0.8, alpha=0.85)
        for x, y in zip(xs, ys):
            ax.annotate(label, (x, y), fontsize=5.5, alpha=0.7, ha="center", va="bottom")

    ax.set_title(title, fontsize=14, fontweight="bold")
    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=6.5, ncol=1)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, filename), dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  ✓ Saved {filename}")

    # Silhouette score (if >= 2 labels with >= 2 samples each)
    label_counts = pd.Series(labels).value_counts()
    valid = label_counts[label_counts >= 2].index.tolist()
    if len(valid) >= 2:
        mask = [l in valid for l in labels]
        filtered_embs = embeddings[[i for i, m in enumerate(mask) if m]]
        filtered_labels = [l for l, m in zip(labels, mask) if m]
        score = silhouette_score(filtered_embs, filtered_labels, metric="cosine")
        print(f"  📊 Silhouette Score (cosine, frames with ≥2 samples): {score:.4f}")
        return score
    return None


# ═══════════════════════════════════════════════════════════
# PART 2: Token-Level Embedding UMAP
# ═══════════════════════════════════════════════════════════

def extract_token_embeddings(df: pd.DataFrame, tokenizer, model, device) -> tuple[np.ndarray, list[str], list[str]]:
    """Extract contextualized embeddings for frame-evoking tokens."""
    all_embeddings = []
    all_labels = []
    all_token_texts = []

    model.eval()
    with torch.no_grad():
        for _, row in df.iterrows():
            text = row["text"]
            inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs)
            hidden = outputs.last_hidden_state.squeeze(0)  # (seq_len, 768)

            # For each frame-evoking token, find its subword indices
            for token_str in row["tokens"]:
                # Tokenize just the frame token to get its subword tokens
                tok_ids = tokenizer.encode(token_str, add_special_tokens=False)
                # Find these token IDs in the full sequence
                input_ids = inputs["input_ids"].squeeze(0).tolist()

                # Sliding window search for subword match
                found = False
                for start_i in range(len(input_ids) - len(tok_ids) + 1):
                    if input_ids[start_i:start_i + len(tok_ids)] == tok_ids:
                        # Average the subword embeddings
                        token_emb = hidden[start_i:start_i + len(tok_ids)].mean(dim=0).cpu().numpy()
                        all_embeddings.append(token_emb)
                        all_labels.append(row["core_frame"])
                        all_token_texts.append(token_str)
                        found = True
                        break

                if not found:
                    # Fallback: try with leading space (RoBERTa tokenization quirk)
                    tok_ids2 = tokenizer.encode(" " + token_str, add_special_tokens=False)
                    for start_i in range(len(input_ids) - len(tok_ids2) + 1):
                        if input_ids[start_i:start_i + len(tok_ids2)] == tok_ids2:
                            token_emb = hidden[start_i:start_i + len(tok_ids2)].mean(dim=0).cpu().numpy()
                            all_embeddings.append(token_emb)
                            all_labels.append(row["core_frame"])
                            all_token_texts.append(token_str)
                            break

    return np.array(all_embeddings), all_labels, all_token_texts


def plot_umap_tokens(embeddings: np.ndarray, labels: list[str], token_texts: list[str], filename: str):
    """UMAP visualization for frame-evoking token embeddings."""
    try:
        import umap
    except ImportError:
        import umap.umap_ as umap

    n = len(embeddings)
    if n < 5:
        print("  ⚠ Not enough token embeddings to plot UMAP")
        return

    reducer = umap.UMAP(n_neighbors=min(10, n - 1), min_dist=0.2, random_state=42, metric="cosine")
    coords = reducer.fit_transform(embeddings)

    unique_labels = sorted(set(labels))
    color_map = {label: cm.tab20(i / max(len(unique_labels), 1)) for i, label in enumerate(unique_labels)}

    fig, ax = plt.subplots(figsize=(18, 14))
    for label in unique_labels:
        mask = [l == label for l in labels]
        idxs = [i for i, m in enumerate(mask) if m]
        xs = coords[idxs, 0]
        ys = coords[idxs, 1]
        ax.scatter(xs, ys, c=[color_map[label]], label=label, s=80, edgecolors="white", linewidth=0.5, alpha=0.8)
        for x, y, idx in zip(xs, ys, idxs):
            ax.annotate(token_texts[idx], (x, y), fontsize=5, alpha=0.65, ha="center", va="bottom")

    ax.set_title("Token-Level Embeddings (Frame-Evoking Tokens) — Base ClimateBERT", fontsize=13, fontweight="bold")
    ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", fontsize=6, ncol=1)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, filename), dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  ✓ Saved {filename}")


# ═══════════════════════════════════════════════════════════
# PART 3: Frame Similarity Heatmap
# ═══════════════════════════════════════════════════════════

def plot_frame_similarity(embeddings: np.ndarray, labels: list[str], filename: str):
    """Compute per-frame centroid embeddings and plot cosine similarity heatmap."""
    unique_frames = sorted(set(labels))
    centroids = []
    for frame in unique_frames:
        mask = [l == frame for l in labels]
        frame_embs = embeddings[[i for i, m in enumerate(mask) if m]]
        centroids.append(frame_embs.mean(axis=0))
    centroids = np.array(centroids)

    sim_matrix = cosine_similarity(centroids)
    df_sim = pd.DataFrame(sim_matrix, index=unique_frames, columns=unique_frames)

    fig, ax = plt.subplots(figsize=(16, 14))
    sns.heatmap(df_sim, annot=True, fmt=".2f", cmap="RdYlBu_r", center=0.5,
                xticklabels=True, yticklabels=True, ax=ax,
                annot_kws={"fontsize": 5.5}, linewidths=0.3)
    ax.set_title("Frame-to-Frame Cosine Similarity (Paragraph Centroids) — Base ClimateBERT", fontsize=12, fontweight="bold")
    ax.tick_params(labelsize=6.5)
    plt.xticks(rotation=45, ha="right")
    plt.yticks(rotation=0)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, filename), dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  ✓ Saved {filename}")


# ═══════════════════════════════════════════════════════════
# PART 4: MLM Probing
# ═══════════════════════════════════════════════════════════

def mlm_probing(df: pd.DataFrame, tokenizer, mlm_model, device, n_examples: int = 5, top_k: int = 5):
    """Mask frame-evoking tokens and check what the model predicts."""
    results = []
    mlm_model.eval()
    mask_token = tokenizer.mask_token

    for _, row in df.head(n_examples).iterrows():
        text = row["text"]
        for token_str in row["tokens"][:2]:  # first 2 tokens per paragraph
            # Replace the token with <mask>
            pattern = re.compile(re.escape(token_str), re.IGNORECASE)
            masked_text = pattern.sub(mask_token, text, count=1)
            if mask_token not in masked_text:
                continue  # token not found in text

            inputs = tokenizer(masked_text, return_tensors="pt", truncation=True, max_length=512).to(device)
            mask_positions = (inputs["input_ids"] == tokenizer.mask_token_id).nonzero(as_tuple=True)

            if len(mask_positions[0]) == 0:
                continue

            with torch.no_grad():
                outputs = mlm_model(**inputs)
                logits = outputs.logits

            # Get predictions for first mask position
            mask_idx = mask_positions[1][0].item()
            probs = torch.softmax(logits[0, mask_idx], dim=-1)
            top_probs, top_ids = probs.topk(top_k)

            predictions = [(tokenizer.decode([tid]).strip(), f"{p:.4f}") for tid, p in zip(top_ids.tolist(), top_probs.tolist())]

            results.append({
                "frame": row["core_frame"],
                "masked_token": token_str,
                "predictions": predictions,
            })

    return results


# ═══════════════════════════════════════════════════════════
# PART 5: Attention Analysis
# ═══════════════════════════════════════════════════════════

def attention_analysis(df: pd.DataFrame, tokenizer, model, device, n_examples: int = 3):
    """Analyze attention weights: do frame-evoking tokens receive more attention?"""
    results = []
    model.eval()

    for _, row in df.head(n_examples).iterrows():
        text = row["text"]
        inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512).to(device)

        with torch.no_grad():
            outputs = model(**inputs, output_attentions=True)

        # Average attention across all heads and layers → (seq_len, seq_len)
        attentions = torch.stack(outputs.attentions)  # (n_layers, batch, n_heads, seq, seq)
        mean_attn = attentions.mean(dim=(0, 1, 2)).cpu().numpy()  # (seq, seq)

        # Attention received by each token (column-wise sum)
        attn_received = mean_attn.sum(axis=0)  # how much attention each token receives
        attn_received = attn_received / attn_received.sum()  # normalize

        input_ids = inputs["input_ids"].squeeze(0).tolist()
        tokens_decoded = [tokenizer.decode([tid]).strip() for tid in input_ids]

        # Find frame-evoking token indices
        frame_attn_scores = {}
        for token_str in row["tokens"]:
            tok_ids = tokenizer.encode(" " + token_str, add_special_tokens=False)
            for start_i in range(len(input_ids) - len(tok_ids) + 1):
                if input_ids[start_i:start_i + len(tok_ids)] == tok_ids:
                    avg_attn = attn_received[start_i:start_i + len(tok_ids)].mean()
                    frame_attn_scores[token_str] = float(avg_attn)
                    break

        # Average attention for non-frame tokens
        all_frame_indices = set()
        for token_str in row["tokens"]:
            tok_ids = tokenizer.encode(" " + token_str, add_special_tokens=False)
            for start_i in range(len(input_ids) - len(tok_ids) + 1):
                if input_ids[start_i:start_i + len(tok_ids)] == tok_ids:
                    all_frame_indices.update(range(start_i, start_i + len(tok_ids)))
                    break

        non_frame_indices = [i for i in range(1, len(input_ids) - 1) if i not in all_frame_indices]  # skip special tokens
        avg_non_frame = attn_received[non_frame_indices].mean() if non_frame_indices else 0

        results.append({
            "frame": row["core_frame"],
            "text_preview": text[:80] + "...",
            "frame_token_attention": frame_attn_scores,
            "avg_non_frame_attention": float(avg_non_frame),
            "attn_ratio": np.mean(list(frame_attn_scores.values())) / max(avg_non_frame, 1e-8) if frame_attn_scores else 0,
        })

    return results


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("  ClimateBERT Base Model POC — Capabilities Analysis")
    print("=" * 70)

    # --- Load Data ---
    print("\n📂 Loading dataset...")
    df = load_data(DATA_PATH)
    print(f"   {len(df)} paragraphs loaded")
    print(f"   {len(df['core_frame'].unique())} unique core frames")
    all_periph = set()
    for pf in df["peripheral_frames"]:
        all_periph.update(pf)
    print(f"   {len(all_periph)} unique peripheral frames")
    total_tokens = sum(len(t) for t in df["tokens"])
    print(f"   {total_tokens} total frame-evoking token instances")

    # --- Load Model ---
    print(f"\n🤖 Loading model: {MODEL_NAME}")
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"   Device: {device}")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    model = AutoModel.from_pretrained(MODEL_NAME, output_attentions=True).to(device)
    mlm_model = AutoModelForMaskedLM.from_pretrained(MODEL_NAME).to(device)
    print("   ✓ Model loaded")

    # =============================================================
    # MODULE 1: Paragraph-Level Embeddings
    # =============================================================
    print("\n" + "─" * 50)
    print("📊 Module 1: Paragraph-Level Embedding Analysis")
    print("─" * 50)

    print("  Extracting [CLS] embeddings...")
    cls_embs = extract_paragraph_embeddings(df["text"].tolist(), tokenizer, model, device)
    print(f"  Shape: {cls_embs.shape}")

    print("  Extracting mean-pooled embeddings...")
    mean_embs = extract_mean_pooled_embeddings(df["text"].tolist(), tokenizer, model, device)
    print(f"  Shape: {mean_embs.shape}")

    # UMAP with Core Frame labels
    print("\n  Plotting UMAP (Core Frame labels, CLS embeddings)...")
    sil_cls = plot_umap_paragraphs(cls_embs, df["core_frame"].tolist(),
                                    "Paragraph Embeddings [CLS] — Core Frame — Base ClimateBERT",
                                    "paragraph_umap_cls_core.png")

    print("  Plotting UMAP (Core Frame labels, Mean-Pooled)...")
    sil_mean = plot_umap_paragraphs(mean_embs, df["core_frame"].tolist(),
                                     "Paragraph Embeddings [Mean-Pool] — Core Frame — Base ClimateBERT",
                                     "paragraph_umap_mean_core.png")

    # =============================================================
    # MODULE 2: Token-Level Embeddings
    # =============================================================
    print("\n" + "─" * 50)
    print("📊 Module 2: Token-Level Embedding Analysis")
    print("─" * 50)

    print("  Extracting frame-evoking token embeddings...")
    tok_embs, tok_labels, tok_texts = extract_token_embeddings(df, tokenizer, model, device)
    print(f"  Extracted {len(tok_embs)} token embeddings")

    if len(tok_embs) > 0:
        print("  Plotting UMAP (token embeddings, Core Frame colors)...")
        plot_umap_tokens(tok_embs, tok_labels, tok_texts, "token_umap_core.png")

    # =============================================================
    # MODULE 3: Frame Similarity Heatmap
    # =============================================================
    print("\n" + "─" * 50)
    print("📊 Module 3: Frame-to-Frame Similarity")
    print("─" * 50)

    print("  Computing frame centroids and similarity matrix...")
    plot_frame_similarity(mean_embs, df["core_frame"].tolist(), "frame_similarity_heatmap.png")

    # =============================================================
    # MODULE 4: MLM Probing
    # =============================================================
    print("\n" + "─" * 50)
    print("📊 Module 4: MLM Probing (Domain Understanding)")
    print("─" * 50)

    print("  Running MLM probing on 5 sample paragraphs...")
    mlm_results = mlm_probing(df, tokenizer, mlm_model, device, n_examples=5, top_k=5)

    # Save MLM results
    mlm_report_lines = []
    for r in mlm_results:
        mlm_report_lines.append(f"\n  Frame: {r['frame']}")
        mlm_report_lines.append(f"  Masked token: \"{r['masked_token']}\"")
        mlm_report_lines.append(f"  Top predictions: {r['predictions']}")
    mlm_report = "\n".join(mlm_report_lines)
    print(mlm_report)

    with open(os.path.join(OUTPUT_DIR, "mlm_probing_results.txt"), "w", encoding="utf-8") as f:
        f.write("MLM Probing Results — ClimateBERT Base Model\n")
        f.write("=" * 50 + "\n")
        f.write(mlm_report)
    print("  ✓ Saved mlm_probing_results.txt")

    # =============================================================
    # MODULE 5: Attention Analysis
    # =============================================================
    print("\n" + "─" * 50)
    print("📊 Module 5: Attention Weight Analysis")
    print("─" * 50)

    print("  Analyzing attention patterns on 3 sample paragraphs...")
    attn_results = attention_analysis(df, tokenizer, model, device, n_examples=3)

    attn_report_lines = []
    for r in attn_results:
        attn_report_lines.append(f"\n  Frame: {r['frame']}")
        attn_report_lines.append(f"  Text: {r['text_preview']}")
        attn_report_lines.append(f"  Frame token attention scores:")
        for tok, score in r["frame_token_attention"].items():
            attn_report_lines.append(f"    \"{tok}\": {score:.6f}")
        attn_report_lines.append(f"  Avg non-frame token attention: {r['avg_non_frame_attention']:.6f}")
        attn_report_lines.append(f"  Attention ratio (frame/non-frame): {r['attn_ratio']:.2f}x")
    attn_report = "\n".join(attn_report_lines)
    print(attn_report)

    with open(os.path.join(OUTPUT_DIR, "attention_analysis_results.txt"), "w", encoding="utf-8") as f:
        f.write("Attention Analysis Results — ClimateBERT Base Model\n")
        f.write("=" * 50 + "\n")
        f.write(attn_report)
    print("  ✓ Saved attention_analysis_results.txt")

    # =============================================================
    # SUMMARY
    # =============================================================
    print("\n" + "=" * 70)
    print("  📋 POC Summary")
    print("=" * 70)
    print(f"  Dataset: {len(df)} paragraphs, {len(df['core_frame'].unique())} core frames, {total_tokens} frame tokens")
    print(f"  Model: {MODEL_NAME} (DistilRoBERTa, 82.4M params)")
    print(f"  Device: {device}")
    if sil_cls is not None:
        print(f"  Silhouette (CLS): {sil_cls:.4f}")
    if sil_mean is not None:
        print(f"  Silhouette (Mean-Pool): {sil_mean:.4f}")
    print(f"  Token embeddings extracted: {len(tok_embs)}")
    print(f"  MLM probes: {len(mlm_results)}")
    print(f"  Attention analyses: {len(attn_results)}")
    print(f"\n  All outputs saved to: {OUTPUT_DIR}")
    print("=" * 70)


if __name__ == "__main__":
    main()
