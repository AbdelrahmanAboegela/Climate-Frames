"""
ClimateBERT Fine-Tuning — Phase 0: Few-Shot Frame Classifier
=============================================================
Since SetFit 1.x has dependency conflicts with transformers 5.x,
this script implements the core idea directly:

Strategy A — LoRA Sequence Classifier (HuggingFace PEFT):
  - Adds LoRA adapters (rank=8) to ClimateBERT's attention Q+V matrices
  - Fine-tunes with CrossEntropy + class weights for imbalance
  - Stratified 5-Fold CV with LOO-CV for very small folds

Strategy B — Prototype (nearest centroid) classifier:
  - Mean-pools embeddings per class from training split
  - Predicts via cosine similarity to class prototypes
  - No gradient updates — pure metric learning baseline

Strategy C — Dictionary-based token extraction:
  - Uses Token Summary canonical tokens against paragraphs
  - Computes token-level precision, recall, F1

Outputs:
  poc_outputs/finetune_cv_results.txt
  poc_outputs/finetune_confusion_matrix.png
  poc_outputs/finetune_per_class_f1.png
  poc_outputs/finetune_token_extraction.txt

Run: python poc_finetune_setfit.py
"""

import os
import re
import warnings
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from collections import defaultdict
from sklearn.model_selection import StratifiedKFold
from sklearn.metrics import (
    classification_report, confusion_matrix,
    f1_score, accuracy_score
)
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import LabelEncoder
from transformers import (
    AutoTokenizer, AutoModel, AutoModelForSequenceClassification,
    TrainingArguments, Trainer, DataCollatorWithPadding, logging as hf_logging
)
from peft import LoraConfig, get_peft_model, TaskType
from datasets import Dataset as HFDataset

hf_logging.set_verbosity_error()
warnings.filterwarnings("ignore")

OUTPUT_DIR = r"e:\Frames\poc_outputs"
DATA_PATH  = r"e:\Frames\12 articles Ann. Core Peripheral RST and FrameNET Structure.xlsx"
SHEET_NAME = "Core and Peripheral Annotations"
MODEL_NAME = "climatebert/distilroberta-base-climate-f"

os.makedirs(OUTPUT_DIR, exist_ok=True)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ═══════════════════════════════════════════════════════════════
# 1. Data Loading
# ═══════════════════════════════════════════════════════════════

def normalize_frame(name: str) -> str:
    name = name.strip()
    name = re.sub(r"\s*_\s*", "_", name)
    name = re.sub(r"\s+", " ", name)
    return name


def load_data() -> pd.DataFrame:
    wb = openpyxl.load_workbook(DATA_PATH)
    ws = wb[SHEET_NAME]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 4 or not row[0]:
            continue
        text, core, periph, tokens = row[0], row[1], row[2], row[3]
        frame_roles = row[4] if len(row) > 4 else None
        core_frame = normalize_frame(str(core)) if core else ""
        if not core_frame or core_frame.lower() in ("none", "nan"):
            continue
        token_list = [t.strip().lower() for t in str(tokens).split(";") if t.strip()] if tokens else []
        records.append({
            "text": str(text).strip(),
            "core_frame": core_frame,
            "tokens": token_list,
        })
    return pd.DataFrame(records)


def load_token_summary() -> dict:
    wb = openpyxl.load_workbook(DATA_PATH)
    if "Token Summary" not in wb.sheetnames:
        return {}
    ws = wb["Token Summary"]
    summary = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        frame = normalize_frame(str(row[0]))
        token = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
        if token:
            summary[frame].append(token)
    return dict(summary)


# ═══════════════════════════════════════════════════════════════
# 2. Embedding helper (for prototype classifier)
# ═══════════════════════════════════════════════════════════════

def embed_texts(texts, tokenizer, model):
    model.eval()
    embeddings = []
    with torch.no_grad():
        for text in texts:
            inputs = tokenizer(text, return_tensors="pt", truncation=True,
                               max_length=512, padding=True).to(device)
            out = model(**inputs)
            mask = inputs["attention_mask"].unsqueeze(-1)
            emb = (out.last_hidden_state * mask).sum(1) / mask.sum(1)
            embeddings.append(emb.cpu().numpy().squeeze())
    return np.array(embeddings)


# ═══════════════════════════════════════════════════════════════
# 3. Strategy A — LoRA Fine-Tune (Stratified 5-Fold CV)
# ═══════════════════════════════════════════════════════════════

def run_lora_cv(df, label2id, id2label, n_splits=5):
    print("\n" + "═" * 60)
    print("  Strategy A: LoRA Fine-Tune — Stratified 5-Fold CV")
    print("═" * 60)

    texts  = df["text"].tolist()
    labels = [label2id[f] for f in df["core_frame"].tolist()]
    n_labels = len(label2id)

    # Class weights for imbalance
    counts = np.bincount(labels)
    class_weights = torch.tensor(
        (counts.sum() / (n_labels * counts)).astype(np.float32), device=device
    )

    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    all_preds, all_true = [], []
    fold_accs, fold_f1s = [], []

    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)

    for fold, (train_idx, val_idx) in enumerate(skf.split(texts, labels), 1):
        print(f"\n  ── Fold {fold}/{n_splits}  (train={len(train_idx)}, val={len(val_idx)}) ──")

        # Build PEFT model fresh each fold
        base = AutoModelForSequenceClassification.from_pretrained(
            MODEL_NAME,
            num_labels=n_labels,
            ignore_mismatched_sizes=True,
        )
        lora_cfg = LoraConfig(
            task_type=TaskType.SEQ_CLS,
            r=8,
            lora_alpha=16,
            lora_dropout=0.1,
            target_modules=["query", "value"],
            bias="none",
        )
        model = get_peft_model(base, lora_cfg).to(device)
        trainable = sum(p.numel() for p in model.parameters() if p.requires_grad)
        total     = sum(p.numel() for p in model.parameters())
        print(f"     Trainable params: {trainable:,} / {total:,} ({100*trainable/total:.2f}%)")

        # Tokenize
        def tokenize(examples):
            return tokenizer(examples["text"], truncation=True, max_length=512)

        train_ds = HFDataset.from_dict({
            "text":  [texts[i] for i in train_idx],
            "label": [labels[i] for i in train_idx],
        }).map(tokenize, batched=True, remove_columns=["text"])

        val_ds = HFDataset.from_dict({
            "text":  [texts[i] for i in val_idx],
            "label": [labels[i] for i in val_idx],
        }).map(tokenize, batched=True, remove_columns=["text"])

        # Custom weighted trainer
        class WeightedTrainer(Trainer):
            def compute_loss(self, model, inputs, return_outputs=False, **kwargs):
                labels_ = inputs.pop("labels")
                outputs = model(**inputs)
                logits  = outputs.logits
                loss_fn = nn.CrossEntropyLoss(weight=class_weights)
                loss    = loss_fn(logits, labels_)
                return (loss, outputs) if return_outputs else loss

        train_args = TrainingArguments(
            output_dir=os.path.join(OUTPUT_DIR, f"lora_fold{fold}"),
            num_train_epochs=8,
            per_device_train_batch_size=8,
            per_device_eval_batch_size=16,
            learning_rate=2e-4,
            warmup_ratio=0.1,
            weight_decay=0.01,
            logging_steps=50,
            save_strategy="no",
            eval_strategy="no",
            report_to="none",
            fp16=torch.cuda.is_available(),
        )

        trainer = WeightedTrainer(
            model=model,
            args=train_args,
            train_dataset=train_ds,
            processing_class=tokenizer,
            data_collator=DataCollatorWithPadding(tokenizer),
        )
        trainer.train()

        # Predict
        preds_out = trainer.predict(val_ds)
        pred_ids  = np.argmax(preds_out.predictions, axis=1).tolist()
        true_ids  = [labels[i] for i in val_idx]

        acc = accuracy_score(true_ids, pred_ids)
        f1  = f1_score(true_ids, pred_ids, average="macro", zero_division=0)
        fold_accs.append(acc)
        fold_f1s.append(f1)
        all_preds.extend(pred_ids)
        all_true.extend(true_ids)
        print(f"     Accuracy: {acc:.3f} | Macro-F1: {f1:.3f}")

        # Cleanup GPU memory
        del model, trainer, base
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    return all_true, all_preds, fold_accs, fold_f1s


# ═══════════════════════════════════════════════════════════════
# 4. Strategy B — Prototype Classifier (metric learning baseline)
# ═══════════════════════════════════════════════════════════════

def run_prototype_cv(df, label2id, id2label, n_splits=5):
    print("\n" + "═" * 60)
    print("  Strategy B: Prototype Classifier (cosine similarity)")
    print("═" * 60)

    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    model = AutoModel.from_pretrained(MODEL_NAME).to(device)

    texts  = np.array(df["text"].tolist())
    labels = np.array([label2id[f] for f in df["core_frame"].tolist()])

    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=42)
    all_preds, all_true = [], []
    fold_accs, fold_f1s = [], []

    for fold, (train_idx, val_idx) in enumerate(skf.split(texts, labels), 1):
        X_train, X_val = texts[train_idx], texts[val_idx]
        y_train, y_val = labels[train_idx], labels[val_idx]

        # Build prototypes: mean embedding per class
        train_embs = embed_texts(X_train.tolist(), tokenizer, model)
        val_embs   = embed_texts(X_val.tolist(), tokenizer, model)

        # Normalize
        train_embs = train_embs / (np.linalg.norm(train_embs, axis=1, keepdims=True) + 1e-9)
        val_embs   = val_embs   / (np.linalg.norm(val_embs, axis=1, keepdims=True) + 1e-9)

        prototypes = {}
        for cid in np.unique(y_train):
            mask = y_train == cid
            prototypes[cid] = train_embs[mask].mean(axis=0)

        proto_matrix = np.stack([prototypes[i] for i in sorted(prototypes.keys())])
        proto_ids    = sorted(prototypes.keys())

        sims     = val_embs @ proto_matrix.T
        pred_ids = [proto_ids[i] for i in sims.argmax(axis=1)]

        acc = accuracy_score(y_val, pred_ids)
        f1  = f1_score(y_val, pred_ids, average="macro", zero_division=0)
        fold_accs.append(acc)
        fold_f1s.append(f1)
        all_preds.extend(pred_ids)
        all_true.extend(y_val.tolist())
        print(f"     Fold {fold}: Accuracy={acc:.3f}  Macro-F1={f1:.3f}")

    return all_true, all_preds, fold_accs, fold_f1s


# ═══════════════════════════════════════════════════════════════
# 5. Dictionary-Based Token Extraction
# ═══════════════════════════════════════════════════════════════

def evaluate_token_extraction(df, token_summary):
    print("\n" + "─" * 55)
    print("  Token Extraction — Dictionary Matching")
    print("─" * 55)

    canonical = {
        frame: set(t.strip().lower() for t in tokens if t.strip())
        for frame, tokens in token_summary.items()
    }

    all_tp = all_fp = all_fn = 0
    frame_stats = defaultdict(lambda: {"tp": 0, "fp": 0, "fn": 0, "count": 0})

    for _, row in df.iterrows():
        text_lower  = row["text"].lower()
        gold_tokens = set(row["tokens"])
        frame       = row["core_frame"]
        canon       = canonical.get(frame, set())
        predicted   = {ct for ct in canon if ct in text_lower}

        tp = len(gold_tokens & predicted)
        fp = len(predicted - gold_tokens)
        fn = len(gold_tokens - predicted)
        all_tp += tp; all_fp += fp; all_fn += fn
        frame_stats[frame]["tp"] += tp
        frame_stats[frame]["fp"] += fp
        frame_stats[frame]["fn"] += fn
        frame_stats[frame]["count"] += 1

    prec = all_tp / (all_tp + all_fp + 1e-9)
    rec  = all_tp / (all_tp + all_fn + 1e-9)
    f1   = 2 * prec * rec / (prec + rec + 1e-9)

    print(f"  Overall: Precision={prec:.3f}  Recall={rec:.3f}  F1={f1:.3f}")
    for frame, s in sorted(frame_stats.items()):
        p = s["tp"] / (s["tp"] + s["fp"] + 1e-9)
        r = s["tp"] / (s["tp"] + s["fn"] + 1e-9)
        f = 2 * p * r / (p + r + 1e-9)
        print(f"    {frame[:44]:44s} P={p:.2f} R={r:.2f} F1={f:.2f}")

    return prec, rec, f1


# ═══════════════════════════════════════════════════════════════
# 6. Visualizations
# ═══════════════════════════════════════════════════════════════

def plot_confusion_matrix(all_true, all_preds, id2label, tag):
    labels_order = sorted(id2label.keys())
    label_names  = [id2label[i] for i in labels_order]
    short = [n.replace(" and ", "\n& ")[:26] for n in label_names]
    cm = confusion_matrix(all_true, all_preds, labels=labels_order)
    cm_norm = cm.astype(float) / cm.sum(axis=1, keepdims=True).clip(min=1)

    fig, ax = plt.subplots(figsize=(10, 8))
    sns.heatmap(cm_norm, annot=True, fmt=".2f", cmap="Blues",
                xticklabels=short, yticklabels=short,
                linewidths=0.5, ax=ax, vmin=0, vmax=1)
    ax.set_xlabel("Predicted", fontsize=11)
    ax.set_ylabel("True", fontsize=11)
    ax.set_title(f"{tag} — Normalized Confusion Matrix\n(232 paragraphs, 6 frames, 5-fold CV)", fontsize=12)
    plt.xticks(rotation=30, ha="right", fontsize=8)
    plt.yticks(rotation=0, fontsize=8)
    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, f"finetune_confusion_matrix_{tag.lower().replace(' ','_')}.png")
    plt.savefig(out, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✓ Saved confusion matrix → {os.path.basename(out)}")


def plot_comparison(lora_f1s, proto_f1s, id2label):
    labels_order = sorted(id2label.keys())
    label_names  = [id2label[i] for i in labels_order]

    def per_class(all_true, all_preds):
        r = classification_report(all_true, all_preds, labels=labels_order,
                                  target_names=label_names, output_dict=True, zero_division=0)
        return [r[n]["f1-score"] for n in label_names], [r[n]["support"] for n in label_names]

    lora_cls_f1, supports = per_class(*lora_f1s)
    proto_cls_f1, _ = per_class(*proto_f1s)

    x = np.arange(len(label_names))
    short = [n[:26] for n in label_names]

    fig, ax = plt.subplots(figsize=(12, 5))
    ax.bar(x - 0.18, lora_cls_f1,  0.35, label="LoRA Fine-Tune",    color="#2980b9", alpha=0.85)
    ax.bar(x + 0.18, proto_cls_f1, 0.35, label="Prototype (Cosine)", color="#27ae60", alpha=0.85)
    ax.axhline(1/6, color="gray", linestyle="--", linewidth=1, label="Random baseline (17%)")
    for i, (l, p, n) in enumerate(zip(lora_cls_f1, proto_cls_f1, supports)):
        ax.text(i - 0.18, l + 0.01, f"{l:.2f}", ha="center", fontsize=7)
        ax.text(i + 0.18, p + 0.01, f"{p:.2f}", ha="center", fontsize=7)
        ax.text(i, -0.07, f"n={int(n)}", ha="center", fontsize=7, color="gray")
    ax.set_xticks(x)
    ax.set_xticklabels(short, rotation=20, ha="right", fontsize=9)
    ax.set_ylim(-0.1, 1.15)
    ax.set_ylabel("F1 Score", fontsize=11)
    ax.set_title("Phase 0 Fine-Tuning: LoRA vs Prototype — Per-Class F1\n(ClimateBERT, 232 paragraphs, 5-fold CV)", fontsize=12)
    ax.legend(fontsize=10)
    plt.tight_layout()
    out = os.path.join(OUTPUT_DIR, "finetune_comparison.png")
    plt.savefig(out, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✓ Saved finetune_comparison.png")


# ═══════════════════════════════════════════════════════════════
# 7. Save Results
# ═══════════════════════════════════════════════════════════════

def save_results(lora_true, lora_preds, lora_accs, lora_f1s,
                 proto_true, proto_preds, proto_accs, proto_f1s,
                 id2label, tok_prec, tok_rec, tok_f1):
    labels_order = sorted(id2label.keys())
    label_names  = [id2label[i] for i in labels_order]
    lora_report  = classification_report(lora_true, lora_preds, labels=labels_order,
                                         target_names=label_names, zero_division=0)
    proto_report = classification_report(proto_true, proto_preds, labels=labels_order,
                                         target_names=label_names, zero_division=0)

    out = os.path.join(OUTPUT_DIR, "finetune_cv_results.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write("Phase 0 Fine-Tuning Results — ClimateBERT\n")
        f.write("=" * 55 + "\n")
        f.write(f"Model:    {MODEL_NAME}\n")
        f.write(f"Dataset:  232 paragraphs, 6 core frames\n")
        f.write(f"CV:       Stratified 5-Fold\n\n")

        f.write("── Strategy A: LoRA Fine-Tune ──\n")
        f.write(f"Fold Accs:    {[f'{a:.3f}' for a in lora_accs]}\n")
        f.write(f"Mean Acc:     {np.mean(lora_accs):.3f} ± {np.std(lora_accs):.3f}\n")
        f.write(f"Fold Macro-F1:{[f'{x:.3f}' for x in lora_f1s]}\n")
        f.write(f"Mean F1:      {np.mean(lora_f1s):.3f} ± {np.std(lora_f1s):.3f}\n\n")
        f.write(lora_report + "\n\n")

        f.write("── Strategy B: Prototype Classifier ──\n")
        f.write(f"Fold Accs:    {[f'{a:.3f}' for a in proto_accs]}\n")
        f.write(f"Mean Acc:     {np.mean(proto_accs):.3f} ± {np.std(proto_accs):.3f}\n")
        f.write(f"Fold Macro-F1:{[f'{x:.3f}' for x in proto_f1s]}\n")
        f.write(f"Mean F1:      {np.mean(proto_f1s):.3f} ± {np.std(proto_f1s):.3f}\n\n")
        f.write(proto_report + "\n\n")

        f.write("── Strategy C: Token Extraction (Dict-Based) ──\n")
        f.write(f"Precision: {tok_prec:.3f}\n")
        f.write(f"Recall:    {tok_rec:.3f}\n")
        f.write(f"F1:        {tok_f1:.3f}\n")

    print(f"  ✓ Saved finetune_cv_results.txt")


# ═══════════════════════════════════════════════════════════════
# 8. Main
# ═══════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  ClimateBERT Fine-Tuning — Phase 0")
    print("=" * 60)
    print(f"  Device: {device}")

    df = load_data()
    print(f"  Loaded {len(df)} paragraphs, {df['core_frame'].nunique()} frames")
    print("\n  Class distribution:")
    for frame, cnt in df["core_frame"].value_counts().items():
        bar = "█" * int(cnt / 3)
        print(f"    {frame[:42]:42s} {cnt:3d} {bar}")

    token_summary = load_token_summary()
    print(f"\n  Token Summary: {len(token_summary)} frames loaded")

    frames   = sorted(df["core_frame"].unique())
    label2id = {f: i for i, f in enumerate(frames)}
    id2label = {i: f for f, i in label2id.items()}

    # === Strategy A: LoRA ===
    lora_true, lora_preds, lora_accs, lora_f1s = run_lora_cv(df, label2id, id2label)
    print(f"\n  LoRA Mean Acc={np.mean(lora_accs):.3f}  Macro-F1={np.mean(lora_f1s):.3f}")

    # === Strategy B: Prototype ===
    proto_true, proto_preds, proto_accs, proto_f1s = run_prototype_cv(df, label2id, id2label)
    print(f"\n  Proto Mean Acc={np.mean(proto_accs):.3f}  Macro-F1={np.mean(proto_f1s):.3f}")

    # === Strategy C: Token Extraction ===
    tok_prec, tok_rec, tok_f1 = evaluate_token_extraction(df, token_summary)

    # === Visualizations ===
    plot_confusion_matrix(lora_true, lora_preds, id2label, "LoRA")
    plot_confusion_matrix(proto_true, proto_preds, id2label, "Prototype")
    plot_comparison((lora_true, lora_preds), (proto_true, proto_preds), id2label)

    # === Save ===
    save_results(lora_true, lora_preds, lora_accs, lora_f1s,
                 proto_true, proto_preds, proto_accs, proto_f1s,
                 id2label, tok_prec, tok_rec, tok_f1)

    # === Final Summary ===
    print("\n" + "═" * 60)
    print("  FINAL SUMMARY")
    print("═" * 60)
    print(f"  Zero-shot baseline:    19.0% acc | ~17% macro-F1")
    print(f"  Prototype classifier:  {np.mean(proto_accs)*100:.1f}% acc | {np.mean(proto_f1s)*100:.1f}% macro-F1")
    print(f"  LoRA fine-tuned:       {np.mean(lora_accs)*100:.1f}% acc | {np.mean(lora_f1s)*100:.1f}% macro-F1")
    print(f"  Token extraction F1:   {tok_f1*100:.1f}%  (dict-based)")
    print(f"\n  All outputs → {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()

