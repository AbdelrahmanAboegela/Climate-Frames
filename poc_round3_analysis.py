"""
ClimateBERT Base Model POC — Round 3: Deep Probing
===================================================
5. Layer-wise embedding quality (which layer has best frame separation?)
6. k-NN probing classifier (LOO-CV: are frames linearly separable?)
7. Sentence-level decomposition (do individual sentences carry frame signal?)
8. ClimateBERT vs vanilla DistilRoBERTa (is domain pre-training worth it?)
9. Gradient saliency (which tokens truly matter to the model?)

All outputs saved to e:/Frames/poc_outputs/
"""

import os, re, warnings
import numpy as np
import pandas as pd
import torch
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from transformers import AutoTokenizer, AutoModel, logging as hf_logging
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.metrics import silhouette_score
from collections import Counter

hf_logging.set_verbosity_error()
warnings.filterwarnings("ignore")

OUTPUT_DIR = r"e:\Frames\poc_outputs"
MODEL_NAME = "climatebert/distilroberta-base-climate-f"
VANILLA_MODEL = "distilroberta-base"
DATA_PATH = r"e:\Frames\3 Articles Samples Annotation 2026.xlsx"


def normalize_frame(name: str) -> str:
    name = name.strip()
    name = re.sub(r"\s*_\s*", "_", name)
    name = re.sub(r"\s+", " ", name)
    return name


def load_data(path: str) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet1"]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        text, core, periph, tokens = row
        if not text:
            continue
        core_frame = normalize_frame(str(core)) if core else ""
        periph_frames = [normalize_frame(p) for p in str(periph).split(",") if p.strip()] if periph else []
        all_frames = [core_frame] + periph_frames
        token_list = [t.strip().lower() for t in str(tokens).split(",") if t.strip()] if tokens else []
        records.append({
            "text": str(text).strip(),
            "core_frame": core_frame,
            "peripheral_frames": periph_frames,
            "all_frames": all_frames,
            "tokens": token_list,
        })
    return pd.DataFrame(records)


def split_sentences(text: str) -> list[str]:
    """Split text into sentences."""
    sents = re.split(r'(?<=[.!?])\s+', text.strip())
    return [s.strip() for s in sents if len(s.strip()) > 10]


# ═══════════════════════════════════════════════════════════
# TEST 5: Layer-wise Embedding Quality
# ═══════════════════════════════════════════════════════════

def layerwise_analysis(df, tokenizer, model, device):
    """Extract embeddings from each transformer layer and compute silhouette scores."""
    print("\n" + "─" * 55)
    print("🔬 Test 5: Layer-wise Embedding Quality")
    print("─" * 55)

    n_layers = model.config.num_hidden_layers  # 6 for DistilRoBERTa
    print(f"  Model has {n_layers} transformer layers + 1 embedding layer")

    # Find frames with ≥2 samples for silhouette
    frame_counts = df["core_frame"].value_counts()
    valid_mask = df["core_frame"].isin(frame_counts[frame_counts >= 2].index)
    valid_indices = df[valid_mask].index.tolist()
    valid_labels = df.loc[valid_mask, "core_frame"].tolist()
    print(f"  Using {len(valid_indices)} paragraphs from {len(set(valid_labels))} frames (≥2 samples)")

    layer_scores_cls = []
    layer_scores_mean = []

    model.eval()
    # Collect all hidden states
    all_hidden_states = [[] for _ in range(n_layers + 1)]  # +1 for embedding layer

    with torch.no_grad():
        for _, row in df.iterrows():
            inputs = tokenizer(row["text"], return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs, output_hidden_states=True)
            hidden_states = outputs.hidden_states  # tuple of (n_layers+1) tensors

            for layer_idx in range(n_layers + 1):
                hs = hidden_states[layer_idx]  # (1, seq_len, hidden_dim)
                cls_emb = hs[0, 0, :].cpu().numpy()
                mask = inputs["attention_mask"][0].unsqueeze(-1).float()
                mean_emb = (hs[0] * mask.to(device)).sum(dim=0) / mask.to(device).sum(dim=0)
                mean_emb = mean_emb.cpu().numpy()
                all_hidden_states[layer_idx].append((cls_emb, mean_emb))

    # Compute silhouette for each layer
    for layer_idx in range(n_layers + 1):
        cls_embs = np.array([all_hidden_states[layer_idx][i][0] for i in valid_indices])
        mean_embs = np.array([all_hidden_states[layer_idx][i][1] for i in valid_indices])

        try:
            sil_cls = silhouette_score(cls_embs, valid_labels, metric="cosine")
        except:
            sil_cls = 0.0
        try:
            sil_mean = silhouette_score(mean_embs, valid_labels, metric="cosine")
        except:
            sil_mean = 0.0

        layer_scores_cls.append(sil_cls)
        layer_scores_mean.append(sil_mean)

        layer_name = "Embedding" if layer_idx == 0 else f"Layer {layer_idx}"
        print(f"  {layer_name:12s}  CLS: {sil_cls:.4f}  Mean: {sil_mean:.4f}")

    # Plot
    fig, ax = plt.subplots(figsize=(10, 6))
    x_labels = ["Embed"] + [f"L{i}" for i in range(1, n_layers + 1)]
    x = range(len(x_labels))

    ax.plot(x, layer_scores_cls, "o-", color="#e74c3c", linewidth=2.5, markersize=10, label="CLS Token", zorder=3)
    ax.plot(x, layer_scores_mean, "s-", color="#3498db", linewidth=2.5, markersize=10, label="Mean-Pooled", zorder=3)

    best_cls = max(range(len(layer_scores_cls)), key=lambda i: layer_scores_cls[i])
    best_mean = max(range(len(layer_scores_mean)), key=lambda i: layer_scores_mean[i])
    ax.annotate(f"Best: {layer_scores_cls[best_cls]:.3f}", (best_cls, layer_scores_cls[best_cls]),
                textcoords="offset points", xytext=(15, 10), fontsize=9, color="#e74c3c", fontweight="bold")
    ax.annotate(f"Best: {layer_scores_mean[best_mean]:.3f}", (best_mean, layer_scores_mean[best_mean]),
                textcoords="offset points", xytext=(15, -15), fontsize=9, color="#3498db", fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(x_labels, fontsize=11)
    ax.set_xlabel("Layer", fontsize=12)
    ax.set_ylabel("Silhouette Score (cosine)", fontsize=12)
    ax.set_title("Layer-wise Frame Separation Quality — ClimateBERT", fontsize=13, fontweight="bold")
    ax.legend(fontsize=11)
    ax.grid(True, alpha=0.3)
    ax.axhline(0, color="gray", linestyle="--", alpha=0.5)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "layerwise_silhouette.png"), dpi=200)
    plt.close()
    print("  ✓ Saved layerwise_silhouette.png")

    return layer_scores_cls, layer_scores_mean, all_hidden_states


# ═══════════════════════════════════════════════════════════
# TEST 6: k-NN Probing Classifier (LOO-CV)
# ═══════════════════════════════════════════════════════════

def knn_probe(df, tokenizer, model, device):
    """Leave-one-out k=1 nearest neighbor classification on frozen embeddings."""
    print("\n" + "─" * 55)
    print("🔬 Test 6: k-NN Probing Classifier (LOO-CV)")
    print("─" * 55)

    # Get embeddings (last layer, mean-pooled)
    embeddings = []
    model.eval()
    with torch.no_grad():
        for _, row in df.iterrows():
            inputs = tokenizer(row["text"], return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs)
            mask = inputs["attention_mask"].unsqueeze(-1).float()
            mean_emb = (outputs.last_hidden_state * mask.to(device)).sum(dim=1) / mask.to(device).sum(dim=1)
            embeddings.append(mean_emb.cpu().numpy().squeeze())
    embeddings = np.array(embeddings)

    sim_matrix = cosine_similarity(embeddings)
    core_frames = df["core_frame"].tolist()
    all_frames_list = df["all_frames"].tolist()

    # --- Test A: k=1 NN on Core Frame ---
    correct_1nn = 0
    correct_3nn = 0
    for i in range(len(df)):
        sims = sim_matrix[i].copy()
        sims[i] = -1  # exclude self
        # k=1
        nn_idx = np.argmax(sims)
        if core_frames[nn_idx] == core_frames[i]:
            correct_1nn += 1
        # k=3: majority vote
        top3_idx = np.argsort(sims)[-3:]
        top3_frames = [core_frames[j] for j in top3_idx]
        vote = Counter(top3_frames).most_common(1)[0][0]
        if vote == core_frames[i]:
            correct_3nn += 1

    acc_1nn = correct_1nn / len(df) * 100
    acc_3nn = correct_3nn / len(df) * 100
    print(f"\n  📊 Core Frame k-NN Classification (LOO-CV):")
    print(f"     1-NN Accuracy: {correct_1nn}/{len(df)} = {acc_1nn:.1f}%")
    print(f"     3-NN Accuracy: {correct_3nn}/{len(df)} = {acc_3nn:.1f}%")

    # --- Test B: Multi-label frame overlap with NN ---
    any_overlap = 0
    for i in range(len(df)):
        sims = sim_matrix[i].copy()
        sims[i] = -1
        nn_idx = np.argmax(sims)
        # Check if ANY frame overlaps between paragraph i and its nearest neighbor
        frames_i = set(all_frames_list[i])
        frames_nn = set(all_frames_list[nn_idx])
        if frames_i & frames_nn:
            any_overlap += 1

    overlap_rate = any_overlap / len(df) * 100
    print(f"\n  📊 Multi-Label Frame Overlap with Nearest Neighbor:")
    print(f"     Any frame overlap: {any_overlap}/{len(df)} = {overlap_rate:.1f}%")

    # --- Test C: Per-frame accuracy for multi-sample frames ---
    frame_counts = Counter(core_frames)
    multi_frames = {f: c for f, c in frame_counts.items() if c >= 2}
    if multi_frames:
        print(f"\n  📊 Per-Frame Accuracy (frames with ≥2 samples):")
        for frame, count in sorted(multi_frames.items()):
            indices = [i for i, f in enumerate(core_frames) if f == frame]
            correct = 0
            for i in indices:
                sims = sim_matrix[i].copy()
                sims[i] = -1
                nn_idx = np.argmax(sims)
                if core_frames[nn_idx] == frame:
                    correct += 1
            print(f"     {frame}: {correct}/{count} = {correct/count*100:.0f}%")

    # Save results
    with open(os.path.join(OUTPUT_DIR, "knn_probe_results.txt"), "w", encoding="utf-8") as f:
        f.write(f"k-NN Probing Classifier — LOO-CV\n{'='*40}\n")
        f.write(f"1-NN Core Frame Accuracy: {acc_1nn:.1f}%\n")
        f.write(f"3-NN Core Frame Accuracy: {acc_3nn:.1f}%\n")
        f.write(f"NN Frame Overlap Rate: {overlap_rate:.1f}%\n")
    print("  ✓ Saved knn_probe_results.txt")

    return acc_1nn, acc_3nn, overlap_rate


# ═══════════════════════════════════════════════════════════
# TEST 7: Sentence-Level Decomposition
# ═══════════════════════════════════════════════════════════

def sentence_decomposition(df, tokenizer, model, device):
    """Split paragraphs into sentences and test frame signal at sentence level."""
    print("\n" + "─" * 55)
    print("🔬 Test 7: Sentence-Level Decomposition")
    print("─" * 55)

    model.eval()

    # First, compute frame centroids from paragraph embeddings
    para_embs = []
    with torch.no_grad():
        for _, row in df.iterrows():
            inputs = tokenizer(row["text"], return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
            outputs = model(**inputs)
            mask = inputs["attention_mask"].unsqueeze(-1).float()
            mean_emb = (outputs.last_hidden_state * mask.to(device)).sum(dim=1) / mask.to(device).sum(dim=1)
            para_embs.append(mean_emb.cpu().numpy().squeeze())
    para_embs = np.array(para_embs)

    # Compute centroids for frames with ≥2 samples
    frame_counts = df["core_frame"].value_counts()
    multi_frames = frame_counts[frame_counts >= 2].index.tolist()
    centroids = {}
    for frame in multi_frames:
        mask = df["core_frame"] == frame
        centroids[frame] = para_embs[mask].mean(axis=0)

    if not centroids:
        print("  ⚠ Not enough multi-sample frames for centroid analysis")
        return

    centroid_names = list(centroids.keys())
    centroid_embs = np.array([centroids[f] for f in centroid_names])

    # Decompose paragraphs into sentences
    total_sentences = 0
    sentence_results = []

    for idx, row in df.iterrows():
        sentences = split_sentences(row["text"])
        if len(sentences) <= 1:
            continue  # Skip single-sentence paragraphs

        total_sentences += len(sentences)

        # Embed each sentence
        sent_embs = []
        with torch.no_grad():
            for sent in sentences:
                inputs = tokenizer(sent, return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
                outputs = model(**inputs)
                mask = inputs["attention_mask"].unsqueeze(-1).float()
                mean_emb = (outputs.last_hidden_state * mask.to(device)).sum(dim=1) / mask.to(device).sum(dim=1)
                sent_embs.append(mean_emb.cpu().numpy().squeeze())
        sent_embs = np.array(sent_embs)

        # Similarity of each sentence to each frame centroid
        sims = cosine_similarity(sent_embs, centroid_embs)

        for s_idx, sent in enumerate(sentences):
            best_frame_idx = np.argmax(sims[s_idx])
            best_frame = centroid_names[best_frame_idx]
            best_sim = sims[s_idx][best_frame_idx]

            # Check sim to parent paragraph's core frame
            if row["core_frame"] in centroid_names:
                core_sim = sims[s_idx][centroid_names.index(row["core_frame"])]
            else:
                core_sim = None

            sentence_results.append({
                "paragraph_idx": idx,
                "sentence_idx": s_idx,
                "sentence": sent[:80],
                "true_core_frame": row["core_frame"],
                "nearest_centroid": best_frame,
                "nearest_sim": best_sim,
                "core_frame_sim": core_sim,
                "match": best_frame == row["core_frame"],
            })

    # Analyze
    matching = sum(1 for r in sentence_results if r["match"])
    total = len(sentence_results)
    match_rate = matching / total * 100 if total else 0

    print(f"\n  📊 Sentence-Level Analysis:")
    print(f"     Multi-sentence paragraphs: {len(set(r['paragraph_idx'] for r in sentence_results))}")
    print(f"     Total sentences analyzed: {total}")
    print(f"     Sentences matching parent's core frame: {matching}/{total} = {match_rate:.1f}%")

    # Per-paragraph breakdown
    print(f"\n  Sentence breakdown (first 5 multi-sentence paragraphs):")
    seen_paras = set()
    count = 0
    for r in sentence_results:
        if r["paragraph_idx"] not in seen_paras:
            seen_paras.add(r["paragraph_idx"])
            count += 1
            if count > 5:
                break
            print(f"\n    ── Paragraph {r['paragraph_idx']} (Core: {r['true_core_frame']})")

        mark = "✓" if r["match"] else "→"
        sim_str = f"(sim={r['core_frame_sim']:.3f})" if r['core_frame_sim'] is not None else ""
        print(f"      {mark} S{r['sentence_idx']}: \"{r['sentence']}...\" → {r['nearest_centroid']} {sim_str}")

    # Save
    with open(os.path.join(OUTPUT_DIR, "sentence_decomposition.txt"), "w", encoding="utf-8") as f:
        f.write(f"Sentence-Level Frame Decomposition\n{'='*40}\n")
        f.write(f"Sentences matching parent core frame: {match_rate:.1f}%\n\n")
        for r in sentence_results:
            mark = "✓" if r["match"] else "✗"
            f.write(f"{mark} P{r['paragraph_idx']} S{r['sentence_idx']}: {r['nearest_centroid']} (sim={r['nearest_sim']:.3f}) | \"{r['sentence']}...\"\n")
    print("  ✓ Saved sentence_decomposition.txt")

    return match_rate


# ═══════════════════════════════════════════════════════════
# TEST 8: ClimateBERT vs Vanilla DistilRoBERTa
# ═══════════════════════════════════════════════════════════

def climate_vs_vanilla(df, device):
    """Compare ClimateBERT vs vanilla DistilRoBERTa on frame separation."""
    print("\n" + "─" * 55)
    print("🔬 Test 8: ClimateBERT vs Vanilla DistilRoBERTa")
    print("─" * 55)

    results = {}

    for model_name, label in [(MODEL_NAME, "ClimateBERT"), (VANILLA_MODEL, "Vanilla DistilRoBERTa")]:
        print(f"\n  Loading {label}...")
        tok = AutoTokenizer.from_pretrained(model_name)
        mdl = AutoModel.from_pretrained(model_name).to(device)
        mdl.eval()

        embeddings = []
        with torch.no_grad():
            for _, row in df.iterrows():
                inputs = tok(row["text"], return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)
                outputs = mdl(**inputs)
                mask = inputs["attention_mask"].unsqueeze(-1).float()
                mean_emb = (outputs.last_hidden_state * mask.to(device)).sum(dim=1) / mask.to(device).sum(dim=1)
                embeddings.append(mean_emb.cpu().numpy().squeeze())
        embeddings = np.array(embeddings)

        # Silhouette
        frame_counts = df["core_frame"].value_counts()
        valid_mask = df["core_frame"].isin(frame_counts[frame_counts >= 2].index)
        valid_labels = df.loc[valid_mask, "core_frame"].tolist()
        valid_embs = embeddings[valid_mask]

        try:
            sil = silhouette_score(valid_embs, valid_labels, metric="cosine")
        except:
            sil = 0.0

        # Intra/inter similarity
        sim_matrix = cosine_similarity(embeddings)
        intra, inter = [], []
        core_frames = df["core_frame"].tolist()
        for i in range(len(df)):
            for j in range(i + 1, len(df)):
                if core_frames[i] == core_frames[j]:
                    intra.append(sim_matrix[i, j])
                else:
                    inter.append(sim_matrix[i, j])

        gap = np.mean(intra) - np.mean(inter) if intra else 0

        # k=1 NN accuracy
        correct = 0
        for i in range(len(df)):
            sims = sim_matrix[i].copy()
            sims[i] = -1
            nn_idx = np.argmax(sims)
            if core_frames[nn_idx] == core_frames[i]:
                correct += 1
        nn_acc = correct / len(df) * 100

        results[label] = {
            "silhouette": sil,
            "intra_mean": np.mean(intra) if intra else 0,
            "inter_mean": np.mean(inter) if inter else 0,
            "gap": gap,
            "nn_acc": nn_acc,
        }
        print(f"  {label}: Silhouette={sil:.4f}, Gap={gap:.4f}, 1-NN={nn_acc:.1f}%")

        del mdl, tok
        torch.cuda.empty_cache()

    # Comparative plot
    labels = list(results.keys())
    metrics = ["silhouette", "gap", "nn_acc"]
    metric_names = ["Silhouette Score", "Intra-Inter Gap", "1-NN Accuracy (%)"]

    fig, axes = plt.subplots(1, 3, figsize=(15, 5))
    colors = ["#e74c3c", "#3498db"]

    for ax, metric, mname in zip(axes, metrics, metric_names):
        vals = [results[l][metric] for l in labels]
        bars = ax.bar(labels, vals, color=colors, edgecolor="white", linewidth=1.5, width=0.5)
        ax.set_title(mname, fontweight="bold", fontsize=11)
        for bar, val in zip(bars, vals):
            fmt = f"{val:.1f}%" if "Acc" in mname else f"{val:.4f}"
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.002,
                    fmt, ha="center", va="bottom", fontweight="bold", fontsize=10)
        ax.set_ylim(bottom=0)

    plt.suptitle("ClimateBERT vs Vanilla DistilRoBERTa — Frame Separation", fontsize=14, fontweight="bold", y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, "climate_vs_vanilla.png"), dpi=200, bbox_inches="tight")
    plt.close()
    print("\n  ✓ Saved climate_vs_vanilla.png")

    return results


# ═══════════════════════════════════════════════════════════
# TEST 9: Gradient Saliency
# ═══════════════════════════════════════════════════════════

def gradient_saliency(df, tokenizer, model, device):
    """Use gradient-based attribution to find which tokens matter most."""
    print("\n" + "─" * 55)
    print("🔬 Test 9: Gradient Saliency Analysis")
    print("─" * 55)

    model.eval()
    # Enable gradients for embeddings
    for param in model.parameters():
        param.requires_grad_(False)

    sample_indices = [0, 2, 4]  # Pick 3 diverse paragraphs
    results_text = []

    for idx in sample_indices:
        row = df.iloc[idx]
        inputs = tokenizer(row["text"], return_tensors="pt", truncation=True, max_length=512, padding=True).to(device)

        # Get the input embeddings
        embeddings = model.get_input_embeddings()
        input_embeds = embeddings(inputs["input_ids"])
        input_embeds = input_embeds.detach().requires_grad_(True)

        # Forward pass with embeddings instead of input_ids
        # We need to remove input_ids and pass inputs_embeds
        model_inputs = {k: v for k, v in inputs.items() if k != "input_ids"}
        model_inputs["inputs_embeds"] = input_embeds

        outputs = model(**model_inputs)

        # Use CLS embedding L2 norm as the scalar target
        cls_emb = outputs.last_hidden_state[0, 0, :]
        target = cls_emb.norm()
        target.backward()

        # Gradient norms per token
        grad_norms = input_embeds.grad[0].norm(dim=-1).detach().cpu().numpy()

        # Map to tokens
        token_ids = inputs["input_ids"][0].cpu().tolist()
        tokens = [tokenizer.decode([tid]).strip() for tid in token_ids]

        # Normalize
        grad_norms = grad_norms / (grad_norms.max() + 1e-8)

        # Get frame tokens for this paragraph
        frame_tokens = set(row["tokens"])

        # Check if high-saliency tokens overlap with frame tokens
        sorted_indices = np.argsort(grad_norms)[::-1]

        results_text.append(f"\n  Paragraph {idx} — Frame: {row['core_frame']}")
        results_text.append(f"  Text: {row['text'][:80]}...")
        results_text.append(f"  Frame tokens: {frame_tokens}")
        results_text.append(f"  Top-10 salient tokens:")

        frame_token_saliencies = []
        non_frame_saliencies = []

        for rank, ti in enumerate(sorted_indices[:15]):
            tok_str = tokens[ti]
            sal = grad_norms[ti]
            is_frame = any(ft in tok_str.lower() or tok_str.lower() in ft for ft in frame_tokens)
            marker = "★" if is_frame else " "
            if rank < 10:
                results_text.append(f"    {marker} #{rank+1}: \"{tok_str}\" = {sal:.3f}")

        # Compute avg saliency for frame vs non-frame tokens
        for ti in range(len(tokens)):
            tok_str = tokens[ti].lower().strip()
            if not tok_str or tok_str in ["<s>", "</s>", "<pad>"]:
                continue
            is_frame = any(ft in tok_str or tok_str in ft for ft in frame_tokens)
            if is_frame:
                frame_token_saliencies.append(grad_norms[ti])
            else:
                non_frame_saliencies.append(grad_norms[ti])

        if frame_token_saliencies and non_frame_saliencies:
            ratio = np.mean(frame_token_saliencies) / (np.mean(non_frame_saliencies) + 1e-8)
            results_text.append(f"  Saliency ratio (frame/non-frame): {ratio:.2f}x")
        else:
            results_text.append(f"  Saliency ratio: N/A (no matched frame tokens)")

    for line in results_text:
        print(line)

    # Save
    with open(os.path.join(OUTPUT_DIR, "gradient_saliency.txt"), "w", encoding="utf-8") as f:
        for line in results_text:
            f.write(line + "\n")
    print("\n  ✓ Saved gradient_saliency.txt")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  ClimateBERT POC — Round 3: Deep Probing")
    print("=" * 60)

    df = load_data(DATA_PATH)
    print(f"\n📂 Loaded {len(df)} paragraphs")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"🤖 Loading ClimateBERT on {device}...")
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME)
    model = AutoModel.from_pretrained(MODEL_NAME).to(device)
    print("   ✓ Ready")

    # Run tests
    layer_cls, layer_mean, _ = layerwise_analysis(df, tokenizer, model, device)
    acc_1nn, acc_3nn, overlap_rate = knn_probe(df, tokenizer, model, device)
    sent_match = sentence_decomposition(df, tokenizer, model, device)
    gradient_saliency(df, tokenizer, model, device)

    # Free memory before loading vanilla model
    del model
    torch.cuda.empty_cache()

    comparison = climate_vs_vanilla(df, device)

    # Summary
    print("\n" + "=" * 60)
    print("  📋 Round 3 Summary")
    print("=" * 60)
    best_layer_cls = max(range(len(layer_cls)), key=lambda i: layer_cls[i])
    best_layer_mean = max(range(len(layer_mean)), key=lambda i: layer_mean[i])
    print(f"  Best layer (CLS):     {'Embed' if best_layer_cls==0 else f'Layer {best_layer_cls}'} (sil={layer_cls[best_layer_cls]:.4f})")
    print(f"  Best layer (Mean):    {'Embed' if best_layer_mean==0 else f'Layer {best_layer_mean}'} (sil={layer_mean[best_layer_mean]:.4f})")
    print(f"  1-NN Accuracy:        {acc_1nn:.1f}%")
    print(f"  3-NN Accuracy:        {acc_3nn:.1f}%")
    print(f"  NN Frame Overlap:     {overlap_rate:.1f}%")
    print(f"  Sentence match rate:  {sent_match:.1f}%" if sent_match else "  Sentence match: N/A")

    cb = comparison.get("ClimateBERT", {})
    vn = comparison.get("Vanilla DistilRoBERTa", {})
    print(f"\n  ClimateBERT vs Vanilla:")
    print(f"    Silhouette:  {cb.get('silhouette',0):.4f} vs {vn.get('silhouette',0):.4f}")
    print(f"    Gap:         {cb.get('gap',0):.4f} vs {vn.get('gap',0):.4f}")
    print(f"    1-NN:        {cb.get('nn_acc',0):.1f}% vs {vn.get('nn_acc',0):.1f}%")

    print(f"\n  All outputs → {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
